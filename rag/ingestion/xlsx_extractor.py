"""
XLSX ingestion: extracts each sheet as structured text blocks.
"""

from pathlib import Path
import openpyxl

from rag.config import get_settings
from rag.utils.logger import get_logger

log = get_logger(__name__)
settings = get_settings()


def _sheet_to_text(sheet) -> str:
    """Convert an openpyxl worksheet to a readable text block."""
    rows = []
    for row in sheet.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        if any(c.strip() for c in cells):
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def extract_xlsx(file_path: str) -> list[dict]:
    """
    Extract text from each sheet of an XLSX file.

    Returns [{text, sheet_name, type}]
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"XLSX not found: {file_path}")

    size_mb = path.stat().st_size / (1024 * 1024)
    if size_mb > settings.max_file_size_mb:
        raise ValueError(f"File too large: {size_mb:.1f} MB")

    log.info("Extracting XLSX", path=str(path), size_mb=round(size_mb, 2))

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    blocks: list[dict] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text = _sheet_to_text(sheet)
        if text.strip():
            blocks.append({
                "text": text,
                "sheet_name": sheet_name,
                "section": sheet_name,
                "is_heading": False,
                "type": "table",
            })

    wb.close()
    log.info("XLSX extraction complete", sheets=len(blocks))
    return blocks
